import pandas as pd
from datetime import datetime
from tkinter import Tk, Toplevel, filedialog, ttk, Label, Entry, StringVar, messagebox
import customtkinter as ctk
import openpyxl
import os
from fpdf import FPDF
from tkinter.colorchooser import askcolor
# Interface principal


root = ctk.CTk()
root.title("Gerenciador de Clientes e Produtos")
root.geometry("800x600")
root.configure(fg_color="#000000")  # Define a cor inicial do fundo


def personalizar_cores():
    def alterar_cor_plano_de_fundo():
        cor = askcolor(title="Selecione a cor do plano de fundo")[1]
        if cor:
            # Alterar cor do plano de fundo (fg_color para CustomTkinter)
            root.configure(fg_color=cor)
            for janela in janelas_abertas:
                janela.configure(fg_color=cor)

    def alterar_cor_botoes():
        cor = askcolor(title="Selecione a cor dos botões")[1]
        if cor:
            for botao in botoes:
                botao.configure(fg_color=cor)

    # Nova janela para personalização
    janela_personalizacao = ctk.CTkToplevel(root)
    janela_personalizacao.title("Personalizar Cores")
    janela_personalizacao.geometry("400x200")
    janela_personalizacao.configure(fg_color=root.cget(
        "fg_color"))  # Cor do root como padrão
    janelas_abertas.append(janela_personalizacao)

    ctk.CTkLabel(janela_personalizacao, text="Personalizar Cores",
                 font=("Arial", 14)).pack(pady=10)

    botao_janela_personalizacao = ctk.CTkButton(
        janela_personalizacao, text="Alterar Cor do Plano de Fundo", command=alterar_cor_plano_de_fundo)
    botao_janela_personalizacao.pack(pady=10)
    botoes.append(botao_janela_personalizacao)

    botao_janela_personalizacao = ctk.CTkButton(
        janela_personalizacao, text="Alterar Cor dos Botões", command=alterar_cor_botoes)
    botao_janela_personalizacao.pack(pady=10)
    botoes.append(botao_janela_personalizacao)


# Listas para rastrear botões e janelas
botoes = []
janelas_abertas = []


# Função para salvar ou atualizar planilha Excel
def salvar_no_excel(novo_cliente, iten_cliente, valor_iten, data_hora):
    df = pd.DataFrame(
        [{"Produto": iten_cliente, "Preço": valor_iten, "Data_Hora": data_hora}])
    arquivo_excel = f"{novo_cliente}.xlsx"
    try:
        df_existente = pd.read_excel(arquivo_excel)
        df = pd.concat([df_existente, df], ignore_index=True)
    except FileNotFoundError:
        print(f"{arquivo_excel} não encontrado. Criando um novo arquivo.")
    with pd.ExcelWriter(arquivo_excel, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Produtos')
        worksheet = writer.sheets['Produtos']
        worksheet.set_column('A:A', 25)
        worksheet.set_column('B:B', 15)
        worksheet.set_column('C:C', 20)
    print(f"Dados salvos com sucesso em {arquivo_excel}!")
    return arquivo_excel

# Função para carregar a planilha e exibir em outra janela


def carregar_planilha():
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel Files", "*.xlsx *.xls")])
    if file_path:
        df = pd.read_excel(file_path)
        mostrar_tabela(df, file_path)

    # Função para mostrar os dados em uma nova janela

# Função para mostrar os dados em uma nova janela


def mostrar_tabela(df, nome_arquivo):
    global tree
    global arquivo_atual

    arquivo_atual = nome_arquivo
    nova_janela = ctk.CTkToplevel(root)  # Substituindo por CTkToplevel
    nova_janela.title(f"Planilha: {os.path.basename(nome_arquivo)}")
    nova_janela.geometry("800x600")
    # Mantém a cor atual do root como padrão
    nova_janela.configure(fg_color=root.cget("fg_color"))

    # Criando o frame para a tabela
    frame = ttk.Frame(nova_janela)
    frame.pack(expand=True, fill="both", padx=10, pady=10)

    # Criando a treeview (tabela)
    tree = ttk.Treeview(frame, columns=df.columns.tolist(), show="headings")
    tree.pack(expand=True, fill="both")

    # Definindo os cabeçalhos
    for col in df.columns:
        tree.heading(col, text=col)
        # Ajuste de largura e centralização
        tree.column(col, width=200, anchor="center")

    # Estilizando a tabela para combinar com a personalização
    # Linhas ímpares com fundo claro
    tree.tag_configure('oddrow', background="#f0f0f0")
    # Linhas pares com fundo mais escuro
    tree.tag_configure('evenrow', background="#e0e0e0")

    # Inserindo os dados da planilha na tabela com a tag para alternar cores
    for idx, row in df.iterrows():
        tags = 'evenrow' if idx % 2 == 0 else 'oddrow'  # Definindo a tag para cada linha
        # Inserindo uma linha com a tag
        tree.insert("", "end", values=row.tolist(), tags=(tags,))

    # Ajustando o estilo do cabeçalho
    style = ttk.Style()
    style.configure("Treeview.Heading", font=(
        "Arial", 20, "bold"), foreground="black")
    style.configure("Treeview", font=("Arial", 18),
                    background="#f9f9f9", rowheight=25)

    def excluir_linha():
        selected_item = tree.selection()
        if selected_item:
            resposta = messagebox.askyesno(
                "Excluir", "Tem certeza que deseja excluir esta linha?")
            if resposta:
                item_values = tree.item(selected_item, 'values')
                df.drop(df[(df['Produto'] == item_values[0]) &
                           (df['Preço'] == float(item_values[1])) &
                           (df['Data_Hora'] == item_values[2])].index, inplace=True)
                df.to_excel(arquivo_atual, index=False)
                tree.delete(selected_item)
                messagebox.showinfo("Sucesso", "Linha excluída com sucesso!")
        else:
            messagebox.showwarning("Aviso", "Nenhuma linha selecionada.")

    def somar_precos():
        try:
            total = df['Preço'].sum()
            messagebox.showinfo(
                "Total", f"O total dos preços é: R$ {total:.2f}")
        except KeyError:
            messagebox.showerror("Erro", "Erro ao calcular a soma dos preços.")

    botao_excluir_linha = ctk.CTkButton(
        nova_janela, text="Excluir Linha", command=excluir_linha)
    botao_excluir_linha.pack(side="left", padx=10, pady=10)
    botoes.append(botao_excluir_linha)

    botao_somar_precos = ctk.CTkButton(
        nova_janela, text="Somar Preços", command=somar_precos)
    botao_somar_precos.pack(side="left", padx=10, pady=10)
    botoes.append(botao_somar_precos)

    botao_converter_pdf = ctk.CTkButton(
        nova_janela, text="Converter para PDF", command=lambda: converter_para_pdf(nome_arquivo))
    botao_converter_pdf.pack(side="left", padx=10, pady=10)
    botoes.append(botao_converter_pdf)

# Função para converter o arquivo Excel em PDF


def converter_para_pdf(file_path):
    try:
        # Carrega os dados do arquivo Excel
        df = pd.read_excel(file_path)

        # Cria o objeto PDF
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Adiciona o título
        pdf.set_font("Arial", style="B", size=16)
        pdf.cell(0, 10, txt="Relatório de Dados", ln=True, align='C')
        pdf.ln(10)

        # Define o tamanho da célula para cabeçalho e dados
        col_widths = [max(pdf.get_string_width(str(col)) + 10, 40)
                      for col in df.columns]
        header_height = 8
        row_height = 7

        # Adiciona o cabeçalho da tabela
        pdf.set_font("Arial", style="B", size=12)
        for i, col in enumerate(df.columns):
            pdf.cell(col_widths[i], header_height, col, border=1, align='C')
        pdf.ln()

        # Adiciona os dados da tabela
        pdf.set_font("Arial", size=10)
        for _, row in df.iterrows():
            for i, value in enumerate(row):
                pdf.cell(col_widths[i], row_height,
                         str(value), border=1, align='C')
            pdf.ln()

        # Salva o PDF na área de trabalho
        desktop = os.path.join(os.path.join(
            os.environ['USERPROFILE']), 'Desktop')
        pdf_file_path = os.path.join(desktop, os.path.basename(
            file_path).replace('.xlsx', '.pdf'))
        pdf.output(pdf_file_path)

        # Exibe mensagem de sucesso
        messagebox.showinfo(
            "Sucesso", f"Arquivo convertido para PDF com sucesso: {pdf_file_path}")

    except Exception as e:
        # Exibe mensagem de erro
        messagebox.showerror("Erro", f"Erro ao converter para PDF: {e}")


# Função para abrir uma janela de exclusão de arquivo Excel
def abrir_tela_excluir_excel():
    janela_exclusao = Toplevel(root)
    janela_exclusao.title("Excluir Arquivo Excel")
    janela_exclusao.geometry("400x250")
    janela_exclusao.configure(bg="#000080")

    ctk.CTkLabel(janela_exclusao,
                 text="Nome do Arquivo (sem extensão):").pack(pady=10)
    nome_arquivo_var = StringVar()
    ctk.CTkEntry(janela_exclusao, textvariable=nome_arquivo_var,
                 font=("Arial", 12)).pack(pady=10)

    def excluir_arquivo():
        nome_arquivo = nome_arquivo_var.get().strip()
        if not nome_arquivo:
            messagebox.showerror(
                "Erro", "O nome do arquivo deve ser fornecido.")
            return

        arquivo_excel = f"{nome_arquivo}.xlsx"
        if os.path.exists(arquivo_excel):
            resposta = messagebox.askyesno(
                "Excluir", f"Tem certeza que deseja excluir o arquivo {arquivo_excel}?")
            if resposta:
                os.remove(arquivo_excel)
                messagebox.showinfo("Sucesso", f"Arquivo {
                                    arquivo_excel} excluído com sucesso!")
        else:
            messagebox.showerror("Erro", f"O arquivo {
                                 arquivo_excel} não foi encontrado.")

    ctk.CTkButton(janela_exclusao, text="Excluir Arquivo",
                  command=excluir_arquivo).pack(pady=10)

# Função para obter os dados do formulário e salvar no Excel


def adicionar_cliente():
    novo_cliente = nome_cliente.get().strip()
    iten_cliente = nome_item.get().strip()
    valor_iten = preco_item.get().strip()

    if not novo_cliente or not iten_cliente or not valor_iten:
        messagebox.showerror("Erro", "Todos os campos devem ser preenchidos.")
        return
    try:
        valor_iten = float(valor_iten)
        if valor_iten <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror("Erro", "O preço deve ser um número positivo.")
        return

    data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    salvar_no_excel(novo_cliente, iten_cliente, valor_iten, data_hora)
    messagebox.showinfo("Sucesso", f"Dados de {
                        novo_cliente} salvos com sucesso!")

# Função para abrir a calculadora


def abrir_calculadora():
    janela_calculadora = ctk.CTkToplevel(root)
    janela_calculadora.title("Calculadora")
    janela_calculadora.geometry("300x400")
    janela_calculadora.configure(bg="#F3F4F6")

    # Entrada da calculadora
    entrada_calc = ctk.CTkEntry(
        janela_calculadora, width=260, font=("Arial", 32), justify="right")
    entrada_calc.grid(row=0, column=0, columnspan=4, pady=10, padx=10)

    # Função para inserir texto na entrada
    def inserir_texto(valor):
        entrada_calc.insert("end", valor)

    # Função para calcular o resultado
    def calcular():
        try:
            resultado = eval(entrada_calc.get())
            entrada_calc.delete(0, "end")
            entrada_calc.insert(0, str(resultado))
        except Exception:
            entrada_calc.delete(0, "end")
            entrada_calc.insert(0, "Erro")

    # Função para limpar a entrada
    def limpar():
        entrada_calc.delete(0, "end")

    # Botões da calculadora
    botoes = [
        ("7", 1, 0), ("8", 1, 1), ("9", 1, 2), ("/", 1, 3),
        ("4", 2, 0), ("5", 2, 1), ("6", 2, 2), ("*", 2, 3),
        ("1", 3, 0), ("2", 3, 1), ("3", 3, 2), ("-", 3, 3),
        ("C", 4, 0), ("0", 4, 1), ("=", 4, 2), ("+", 4, 3),
    ]

    for texto, linha, coluna in botoes:
        if texto == "=":
            ctk.CTkButton(janela_calculadora, text=texto, command=calcular).grid(
                row=linha, column=coluna, padx=5, pady=5, sticky="nsew")
        elif texto == "C":
            ctk.CTkButton(janela_calculadora, text=texto, command=limpar).grid(
                row=linha, column=coluna, padx=5, pady=5, sticky="nsew")
        else:
            ctk.CTkButton(janela_calculadora, text=texto, command=lambda t=texto: inserir_texto(
                t)).grid(row=linha, column=coluna, padx=5, pady=5, sticky="nsew")

    # Ajustar colunas e linhas
    for i in range(5):
        janela_calculadora.grid_rowconfigure(i, weight=2)
        if i < 4:
            janela_calculadora.grid_columnconfigure(i, weight=2)


# Campos do formulário
ctk.CTkLabel(root, text="Nome do Cliente:").pack(pady=5)
nome_cliente = StringVar()
ctk.CTkEntry(root, textvariable=nome_cliente,
             font=("Arial", 22), width=500).pack(pady=5)

ctk.CTkLabel(root, text="Nome do Item:").pack(pady=5)
nome_item = StringVar()
ctk.CTkEntry(root, textvariable=nome_item, font=(
    "Arial", 22), width=500).pack(pady=5)

ctk.CTkLabel(root, text="Preço do Item:").pack(pady=5)
preco_item = StringVar()
ctk.CTkEntry(root, textvariable=preco_item, font=(
    "Arial", 22), width=500).pack(pady=5)


botao_adicionar_cliente = ctk.CTkButton(
    root, text="Adicionar Cliente", command=adicionar_cliente)
botao_adicionar_cliente.pack(pady=10)
botoes.append(botao_adicionar_cliente)

botao_carregar_planilha = ctk.CTkButton(
    root, text="Carregar Planilha", command=carregar_planilha)
botao_carregar_planilha.pack(pady=10)
botoes.append(botao_carregar_planilha)

botao_excluir_excel = ctk.CTkButton(
    root, text="Excluir Arquivo Excel", command=abrir_tela_excluir_excel)
botao_excluir_excel.pack(pady=10)
botoes.append(botao_excluir_excel)

botao_Abrir_calculadora = ctk.CTkButton(
    root, text="Abrir Calculadora", command=abrir_calculadora)
botao_Abrir_calculadora.pack(pady=10)
botoes.append(botao_Abrir_calculadora)


botao_personalizar_cores = ctk.CTkButton(
    root, text="Personalizar Temas", command=personalizar_cores)
botao_personalizar_cores.pack(pady=10)
botoes.append(botao_personalizar_cores)

root.mainloop()
